import streamlit as st
import pdfplumber
import os
import re
import io
from copy import copy
from collections import defaultdict
from openpyxl import load_workbook

# ============================================================
# CONFIGURACIÓN STREAMLIT Y CONSTANTES
# ============================================================
st.set_page_config(page_title="Consolidador Confidelis", layout="wide")

MESES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
    "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
    "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}
MESES_INV = {v: k for k, v in MESES.items()}

# ============================================================
# FUNCIONES DE EXTRACCIÓN (IDÉNTICAS A TU CÓDIGO)
# ============================================================
def extraer_numeros(texto):
    nums = re.findall(r"[\d,]+\.\d+", texto)
    return [float(n.replace(",", "")) for n in nums]

def extraer_todos_numeros(texto):
    nums = re.findall(r"[\d,]+\.?\d*", texto)
    return [float(n.replace(",", "")) for n in nums if n]

def extraer_numero_despues_de(texto, clave):
    idx = texto.find(clave)
    if idx == -1: return None
    sub = texto[idx + len(clave):]
    nums = extraer_numeros(sub)
    return nums[0] if nums else None

def extraer_nombre_cliente(pdf, plataforma):
    texto = pdf.pages[0].extract_text() or ""
    lineas = texto.split("\n")
    if plataforma == "Prestadero":
        for l in lineas:
            if "Periodo:" in l and "Estado de Cuenta" not in l:
                return re.split(r"\s+Periodo:", l)[0].strip().upper()
    else:
        for l in lineas:
            if "Contrato:" in l:
                p = re.split(r"\s+Contrato:", l)[0].strip()
                p = p.replace("PUBLICO EN GENERAL - ", "")
                return p.upper()
    return "DESCONOCIDO"

def detectar_plataforma(texto):
    return "Prestadero" if ("Prestadero" in texto or "PRESTADERO" in texto) else "GBM"

def es_smart_cash(texto):
    for l in texto.split("\n"):
        if "RENTA VARIABLE" in l and "VALORES EN CORTO" not in l:
            nums = extraer_numeros(l)
            if len(nums) >= 2 and nums[1] > 0:
                return False
    return True

def extraer_saldo_anterior(lineas):
    for l in lineas:
        if "VALOR DEL PORTAFOLIO" in l and "TOTAL" not in l:
            nums = extraer_numeros(l)
            if nums: return nums[0]
    return 0.0

def extraer_portafolio_gbm(pdf):
    portafolio = []
    en_desglose = en_acciones = False
    for pag in pdf.pages:
        texto = pag.extract_text() or ""
        for l in texto.split("\n"):
            ls, lu = l.strip(), l.strip().upper()
            if "DESGLOSE DEL PORTAFOLIO" in lu:
                en_desglose = True; continue
            if not en_desglose: continue
            if lu in ("ACCIONES", "ACCIONES DEL SIC"):
                en_acciones = True; continue
            if en_acciones and ("EMISORA" in lu or "MES ANTERIOR" in lu or "EN PR" in lu):
                continue
            if en_acciones and lu.startswith("TOTAL"):
                en_acciones = False; continue
            if lu in ("DESGLOSE DE MOVIMIENTOS", "RENDIMIENTO DEL PORTAFOLIO", "EFECTIVO"):
                en_desglose = en_acciones = False; continue
            if not en_acciones: continue
            
            m = re.match(r"^([A-Z]+(?:\s+\d+)?)\s*\*?\s+", ls)
            if m:
                try:
                    emisora = m.group(1).strip()
                    nums = extraer_todos_numeros(ls[m.end():])
                    if len(nums) >= 8:
                        portafolio.append({
                            "Emisora": emisora,
                            "Títulos Mes Anterior": int(nums[0]),
                            "Títulos Mes Actual": int(nums[1]),
                            "Costo Total": nums[4],
                            "Precio Mercado Mes Anterior": nums[5],
                            "Precio Mercado Mes Actual": nums[6],
                            "Valor a Mercado": nums[7],
                        })
                except (ValueError, IndexError):
                    continue
    return portafolio

def extraer_deuda_gbm(pdf):
    deuda = []
    en_desglose = en_deuda = False
    for pag in pdf.pages:
        texto = pag.extract_text() or ""
        for l in texto.split("\n"):
            ls, lu = l.strip(), l.strip().upper()
            if "DESGLOSE DEL PORTAFOLIO" in lu:
                en_desglose = True; continue
            if not en_desglose: continue
            if "DEUDA EN REPORTO" in lu and "TOTAL" not in lu:
                en_deuda = True; continue
            if en_deuda and ("EMISORA" in lu or "ANTERIOR" in lu): continue
            if en_deuda and lu.startswith("TOTAL"):
                en_deuda = False; continue
            if lu in ("RENTA VARIABLE", "EFECTIVO", "DESGLOSE DE MOVIMIENTOS", "RENDIMIENTO DEL PORTAFOLIO"):
                en_desglose = en_deuda = False; continue
            if not en_deuda: continue
            m = re.match(r"^([A-Z]+\s+\d+)\s+", ls)
            if m:
                try:
                    emisora = m.group(1).strip()
                    nums = extraer_todos_numeros(ls[m.end():])
                    if len(nums) >= 8:
                        deuda.append({
                            "Emisora": emisora,
                            "Títulos Mes Anterior": int(nums[0]),
                            "Títulos Mes Actual": int(nums[1]),
                            "Tasa": nums[2],
                            "Valor del Reporto": nums[7],
                            "% Cartera": nums[9] if len(nums) >= 10 else 0.0,
                        })
                except (ValueError, IndexError):
                    continue
    return deuda

def extraer_movimientos_acciones(pdf):
    movimientos = []
    en_mov = False
    for pag in pdf.pages:
        texto = pag.extract_text() or ""
        for l in texto.split("\n"):
            lu = l.strip().upper()
            if "DESGLOSE DE MOVIMIENTOS" in lu:
                en_mov = True; continue
            if en_mov and lu in ("RENDIMIENTO DEL PORTAFOLIO", "COMPOSICIÓN FISCAL INFORMATIVA"):
                en_mov = False; continue
            if not en_mov: continue
            if "Compra de Acciones" not in l and "Venta de Acciones" not in l: continue
            try:
                ls = l.strip()
                if "Compra de Acciones" in l:
                    op = "Compra"
                    resto = l[l.find("Compra de Acciones.") + len("Compra de Acciones."):].strip()
                else:
                    op = "Venta"
                    resto = l[l.find("Venta de Acciones.") + len("Venta de Acciones."):].strip()
                fm = re.match(r"(\d{2}/\d{2})", ls)
                fecha = fm.group(1) if fm else ""
                em = re.match(r"^([A-Z]+(?:\s+\d+)?)\s+", resto)
                if em:
                    emisora = em.group(1).strip()
                    nums = extraer_todos_numeros(resto[em.end():])
                    movimientos.append({
                        "Fecha": fecha, "Operación": op, "Emisora": emisora,
                        "Títulos": int(nums[0]) if nums else 0,
                        "Precio Unitario": nums[1] if len(nums) >= 2 else 0,
                        "Comisión": nums[2] if len(nums) >= 3 else 0,
                        "Neto": nums[5] if len(nums) >= 6 else 0,
                    })
            except Exception:
                continue
    return movimientos

def extraer_periodo_pdf(pdf, plataforma):
    texto = pdf.pages[0].extract_text() or ""
    if plataforma == "Prestadero":
        m = re.search(r"(\d{4}).(\d{2}).(\d{2})\s+al\s+(\d{4}).(\d{2}).(\d{2})", texto)
        if m:
            anio, mes = int(m.group(4)), int(m.group(5))
            dia_fin = int(m.group(6))
            nombre_mes = MESES_INV.get(mes, str(mes))
            return {"mes": mes, "anio": anio, "mes_nombre": nombre_mes,
                    "periodo": f"01-{dia_fin} {nombre_mes[:3]} DE {anio}"}
    else:
        m = re.search(r"DEL\s+(\d+)\s+AL\s+(\d+)\s+DE\s+(\w+)\s+DE\s+(\d{4})", texto, re.I)
        if m:
            dia_ini, dia_fin = int(m.group(1)), int(m.group(2))
            nombre_mes = m.group(3).upper()
            anio = int(m.group(4))
            mes = MESES.get(nombre_mes, 0)
            return {"mes": mes, "anio": anio, "mes_nombre": nombre_mes,
                    "periodo": f"{dia_ini:02d}-{dia_fin} {nombre_mes[:3]} DE {anio}"}
    return None

# ============================================================
# PASO 1: EXTRAER DATOS DE LOS PDFs SUBIDOS
# ============================================================
def extraer_todos_los_pdfs_en_memoria(archivos_pdf):
    clientes = defaultdict(lambda: {
        "gbm": None, "smart_cash": None, "prestadero": None, "periodo": None
    })
    for archivo in archivos_pdf:
        try:
            with pdfplumber.open(archivo) as pdf:
                texto_p1 = pdf.pages[0].extract_text() or ""
                texto_completo = texto_p1
                for p in pdf.pages[1:]:
                    t = p.extract_text()
                    if t: texto_completo += "\n" + t

                plataforma = detectar_plataforma(texto_completo)
                nombre = extraer_nombre_cliente(pdf, plataforma)
                lineas_p1 = texto_p1.split("\n")
                periodo = extraer_periodo_pdf(pdf, plataforma)

                if periodo and (clientes[nombre]["periodo"] is None
                                or periodo["mes"] > clientes[nombre]["periodo"]["mes"]):
                    clientes[nombre]["periodo"] = periodo

                if plataforma == "Prestadero":
                    abonos = retiros = interes = valor = 0.0
                    for l in lineas_p1:
                        try:
                            if "Abonos:" in l and "Cuenta Abonos:" not in l:
                                v = extraer_numero_despues_de(l, "Abonos:")
                                if v is not None: abonos = v
                            if "Valor de la Cuenta:" in l:
                                v = extraer_numero_despues_de(l, "Valor de la Cuenta:")
                                if v is not None: valor = v
                            if "Interés Recibido" in l or "Interes Recibido" in l:
                                ns = extraer_numeros(l)
                                if ns: interes = ns[0]
                            if "Retiros:" in l and "Detalle" not in l:
                                v = extraer_numero_despues_de(l, "Retiros:")
                                if v is not None: retiros = v
                        except Exception: continue
                    clientes[nombre]["prestadero"] = {
                        "abonos": abonos, "retiros": retiros, "interes": interes, "valor": valor,
                    }
                else:
                    entradas = salidas = valor_total = saldo_ant = 0.0
                    saldo_ant = extraer_saldo_anterior(lineas_p1)
                    for l in lineas_p1:
                        try:
                            if "ENTRADAS DE EFECTIVO" in l:
                                ns = extraer_numeros(l)
                                if ns: entradas = ns[-1]
                            elif "SALIDAS DE EFECTIVO" in l:
                                ns = extraer_numeros(l)
                                if ns: salidas = ns[-1]
                            elif "VALOR DEL PORTAFOLIO" in l and "TOTAL" not in l:
                                ns = extraer_numeros(l)
                                if len(ns) >= 2: valor_total = ns[1]
                                elif ns: valor_total = ns[0]
                        except Exception: continue

                    smart = es_smart_cash(texto_p1)
                    portafolio = deuda = movimientos = []
                    try: deuda = extraer_deuda_gbm(pdf)
                    except Exception: pass
                    if not smart:
                        try: portafolio = extraer_portafolio_gbm(pdf)
                        except Exception: pass
                        try: movimientos = extraer_movimientos_acciones(pdf)
                        except Exception: pass

                    tipo = "smart_cash" if smart else "gbm"
                    clientes[nombre][tipo] = {
                        "entradas": entradas, "salidas": salidas,
                        "valor_total": valor_total, "saldo_anterior": saldo_ant,
                        "portafolio": portafolio, "deuda": deuda, "movimientos": movimientos,
                    }
        except Exception as e:
            st.error(f"Error procesando el PDF: {e}")
    return clientes

# ============================================================
# FUNCIONES DE CONSOLIDACIÓN (TU LÓGICA INTACTA)
# ============================================================
def normalizar_instr(nombre):
    if not nombre or nombre == "-": return ""
    n = str(nombre).upper().strip()
    n = re.sub(r"\n", " ", n)
    n = re.sub(r"\s+", " ", n)
    return n

ALIASES = {
    "FIBRAPL 14": ["FIBRA PL 14", "FIBRA PL14", "FIBRAPL14"],
    "FIHO 12":    ["FIHO12"],
    "FMTY 14":    ["FMTY14"],
    "FUNO 11":    ["FUNO11"],
    "FIBRAMQ 12": ["FIBRAMQ12"],
    "DAHANOS 13": ["DANHOS 13", "DANHOS13", "DAHANOS13"],
    "GLD":        ["GLD (ORO)", "GLD ORO"],
    "SLV":        ["SLV (PLATA)", "SLV PLATA"],
    "MELI":       ["MELIN", "MELI"],
    "NFLX":       ["NFLX"],
    "FCFE 18":    ["FCFE18"],
}

_ALIAS_MAP = {}
for canonical, alts in ALIASES.items():
    group = {normalizar_instr(canonical)} | {normalizar_instr(a) for a in alts}
    for name in group:
        _ALIAS_MAP[name] = group

def instrumentos_coinciden(nombre_pdf, nombre_master):
    np_ = normalizar_instr(nombre_pdf)
    nm_ = normalizar_instr(nombre_master)
    if np_ == nm_: return True
    grupo = _ALIAS_MAP.get(np_, {np_})
    if nm_ in grupo: return True
    if nm_.startswith(np_ + " ") or np_.startswith(nm_ + " "): return True
    return False

def encontrar_fila(ws, texto_buscar, col=1, rango=(1, 50)):
    for r in range(rango[0], rango[1]):
        v = ws.cell(r, col).value
        if v and texto_buscar in str(v).upper():
            return r
    return None

def valor_numerico(v, default=0.0):
    if isinstance(v, (int, float)): return float(v)
    return default

def actualizar_celda(ws, row, col, value, forzar=False):
    from openpyxl.cell.cell import MergedCell
    target_row, target_col = row, col
    merge_range = None
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            target_row, target_col = rng.min_row, rng.min_col
            merge_range = rng
            break
    celda = ws.cell(target_row, target_col)
    if isinstance(celda, MergedCell) and merge_range:
        try: ws.unmerge_cells(str(merge_range))
        except Exception: pass
        celda = ws.cell(target_row, target_col)
    if not forzar and isinstance(celda.value, str) and str(celda.value).startswith("="):
        return
    try:
        celda.value = value
    except AttributeError:
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                try: ws.unmerge_cells(str(rng))
                except Exception: pass
        ws.cell(row, col).value = value

def copiar_formato_fila(ws, fila_origen, fila_destino):
    from openpyxl.cell.cell import MergedCell
    for col in range(1, 16):
        src = ws.cell(fila_origen, col)
        dst = ws.cell(fila_destino, col)
        if isinstance(src, MergedCell) or isinstance(dst, MergedCell): continue
        try:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.number_format = src.number_format
            dst.alignment = copy(src.alignment)
        except AttributeError:
            continue

def insertar_instrumento(ws, fila_totales, datos, fila_ref, periodo):
    ws.insert_rows(fila_totales)
    nueva_fila = fila_totales
    copiar_formato_fila(ws, fila_ref, nueva_fila)
    emisora = datos["emisora"]
    valor = datos["valor_a_mercado"]
    compra = datos["compra_neto"]
    venta = datos["venta_neto"]
    costo_total = datos.get("costo_total")
    b = costo_total if costo_total and costo_total > 0 else (compra if compra > 0 else valor)
    c = valor
    e = c - b
    f = (e / b) if b > 0 else 0.0
    g = c - 0 + venta - compra
    h = (g / b) if b > 0 else 0.0
    ws.cell(nueva_fila, 1).value = emisora
    ws.cell(nueva_fila, 2).value = round(b, 2)
    ws.cell(nueva_fila, 3).value = round(c, 2)
    if periodo:
        ws.cell(nueva_fila, 4).value = f"{periodo['mes_nombre'].lower()} {periodo['anio']}"
    ws.cell(nueva_fila, 5).value = round(e, 2)
    ws.cell(nueva_fila, 6).value = round(f, 10)
    ws.cell(nueva_fila, 7).value = round(g, 2)
    ws.cell(nueva_fila, 8).value = round(h, 10)
    ws.cell(nueva_fila, 9).value = "CONSERVADOR\nESPECULATIVO"
    ws.cell(nueva_fila, 10).value = round(venta, 2)
    ws.cell(nueva_fila, 11).value = round(compra, 2)
    ws.cell(nueva_fila, 14).value = round(c, 2)
    ws.cell(nueva_fila, 15).value = "GBM"
    return fila_totales + 1

def expandir_formulas_totales(ws, fila_totales):
    patron = re.compile(r"(SUM\([A-Z]+)(\d+)(:[A-Z]+)(\d+)(\))")
    nueva_fin = fila_totales - 1
    for col in range(2, 16):
        celda = ws.cell(fila_totales, col)
        val = celda.value
        if not isinstance(val, str) or not val.startswith("="): continue
        nueva_formula = patron.sub(
            lambda m: f"{m.group(1)}{m.group(2)}{m.group(3)}{nueva_fin}{m.group(5)}",
            val,
        )
        if nueva_formula != val:
            celda.value = nueva_formula

def _actualizar_formulas_header(ws, fila_header, fila_totales, ultima_instr):
    patron_sum = re.compile(r"(=SUM\([A-Z]+)(\d+)(:[A-Z]+)(\d+)(\))")
    celda_h9 = ws.cell(9, 8)
    val_h9 = celda_h9.value
    if isinstance(val_h9, str) and val_h9.startswith("=SUM("):
        nueva = patron_sum.sub(
            lambda m: f"{m.group(1)}{m.group(2)}{m.group(3)}{ultima_instr}{m.group(5)}",
            val_h9,
        )
        if nueva != val_h9: celda_h9.value = nueva
    j9 = ws.cell(9, 10).value
    if isinstance(j9, str) and j9.startswith("=C"):
        ws.cell(9, 10).value = f"=C{fila_totales}-I9"

def leer_instrumentos_master(ws, fila_header, fila_totales):
    instrumentos = []
    r = fila_header + 1
    while r < fila_totales:
        nombre = ws.cell(r, 1).value
        if nombre and str(nombre).strip() and str(nombre).strip() != "-":
            fila_fin = r
            for rng in ws.merged_cells.ranges:
                if rng.min_row == r and rng.min_col == 1:
                    fila_fin = rng.max_row
                    break
            instrumentos.append({
                "fila": r, "fila_fin": fila_fin, "nombre": str(nombre).strip(),
                "B": ws.cell(r, 2).value, "C": ws.cell(r, 3).value,
                "D": ws.cell(r, 4).value, "E": ws.cell(r, 5).value,
                "F": ws.cell(r, 6).value, "G": ws.cell(r, 7).value,
                "H": ws.cell(r, 8).value, "I": ws.cell(r, 9).value,
                "J": ws.cell(r, 10).value, "K": ws.cell(r, 11).value,
                "L": ws.cell(r, 12).value, "M": ws.cell(r, 13).value,
                "N": ws.cell(r, 14).value, "O": ws.cell(r, 15).value,
            })
            r = fila_fin + 1
        else:
            r += 1
    return instrumentos

def _mejor_match_deuda(old_c, fuentes_disponibles):
    if not fuentes_disponibles: return None, None
    if old_c <= 0:
        mejor_key = min(fuentes_disponibles, key=lambda k: fuentes_disponibles[k]["valor"])
        return mejor_key, fuentes_disponibles[mejor_key]
    mejor_key = None
    mejor_diff = float("inf")
    for key, fuente in fuentes_disponibles.items():
        diff = abs(fuente["valor"] - old_c)
        if diff < mejor_diff:
            mejor_diff = diff
            mejor_key = key
    if mejor_key is not None:
        return mejor_key, fuentes_disponibles[mejor_key]
    return None, None

def actualizar_hoja(ws, datos, nombre_hoja):
    gbm = datos.get("gbm")
    smart_cash = datos.get("smart_cash")
    prestadero = datos.get("prestadero")
    periodo = datos.get("periodo")

    fila_header = encontrar_fila(ws, "INSTRUMENTO") or 23
    fila_totales = encontrar_fila(ws, "TOTALES", rango=(fila_header, fila_header + 40))
    if not fila_totales: return

    instrumentos = leer_instrumentos_master(ws, fila_header, fila_totales)
    
    pdf_port = {}
    compras_map = defaultdict(float)
    ventas_map = defaultdict(float)

    if gbm:
        for item in gbm.get("portafolio", []):
            key = normalizar_instr(item["Emisora"])
            pdf_port[key] = {"valor": item["Valor a Mercado"], "costo_total": item.get("Costo Total", None)}
        for mov in gbm.get("movimientos", []):
            key = normalizar_instr(mov["Emisora"])
            if mov["Operación"] == "Compra": compras_map[key] += mov["Neto"]
            else: ventas_map[key] += mov["Neto"]

    deuda_gbm_total = 0.0
    if gbm: deuda_gbm_total = sum(d["Valor del Reporto"] for d in gbm.get("deuda", []))

    deuda_sc_total = 0.0
    sc_entradas = sc_salidas = 0.0
    if smart_cash:
        deuda_sc_total = sum(d["Valor del Reporto"] for d in smart_cash.get("deuda", []))
        sc_entradas = smart_cash.get("entradas", 0)
        sc_salidas = smart_cash.get("salidas", 0)

    fuentes_deuda = {}
    if prestadero:
        fuentes_deuda["prestadero"] = {
            "valor": prestadero["valor"], "retiros": prestadero["retiros"],
            "depositos": prestadero["abonos"], "interes": prestadero["interes"], "tipo": "prestadero",
        }
    if smart_cash and deuda_sc_total > 0:
        fuentes_deuda["smart_cash"] = {
            "valor": deuda_sc_total, "retiros": sc_salidas, "depositos": sc_entradas,
            "interes": 0.0, "tipo": "smart_cash",
        }
    if gbm and deuda_gbm_total > 0:
        fuentes_deuda["gbm_deuda"] = {
            "valor": deuda_gbm_total, "retiros": 0.0, "depositos": 0.0,
            "interes": 0.0, "tipo": "gbm_deuda",
        }

    matched_pdf_keys = set()
    deuda_gbm_matched = 0.0
    deuda_sc_matched = False
    efectivo_instr = None

    for instr in instrumentos:
        fila = instr["fila"]
        nom = instr["nombre"]
        nom_n = normalizar_instr(nom)
        old_c = valor_numerico(instr["C"])
        old_b = valor_numerico(instr["B"])

        matched = False
        new_c = None
        new_g = None
        new_j = 0.0
        new_k = 0.0
        pdf_costo_total = None
        es_efectivo = False
        es_prestadero = False

        if "EFECTIVO" in nom_n and "GBM" in nom_n:
            efectivo_instr = instr
            continue

        elif "DEUDA" in nom_n and instr["O"]:
            prov = normalizar_instr(str(instr["O"]))
            fuente_key = None
            if "PRESTADERO" in prov and "prestadero" in fuentes_deuda: fuente_key = "prestadero"
            elif "SMART" in prov and "smart_cash" in fuentes_deuda: fuente_key = "smart_cash"
            elif "GBM" in prov and "gbm_deuda" in fuentes_deuda: fuente_key = "gbm_deuda"

            if fuente_key:
                match_fuente = fuentes_deuda[fuente_key]
                new_c = match_fuente["valor"]
                new_j = match_fuente["retiros"]
                new_k = match_fuente["depositos"]
                es_prestadero = (match_fuente["tipo"] == "prestadero")
                if es_prestadero: new_g = match_fuente["interes"]
                else: new_g = new_c - old_c
                matched = True
                if match_fuente["tipo"] == "gbm_deuda": deuda_gbm_matched += new_c
                elif match_fuente["tipo"] == "smart_cash": deuda_sc_matched = True
                del fuentes_deuda[fuente_key]

        elif "DEUDA" in nom_n:
            pass

        else:
            for pdf_key, pdf_data in pdf_port.items():
                if instrumentos_coinciden(pdf_key, nom_n):
                    new_c = pdf_data["valor"]
                    pdf_costo_total = pdf_data.get("costo_total")
                    compra_neto = compras_map.get(pdf_key, 0)
                    venta_neto = ventas_map.get(pdf_key, 0)
                    new_j = venta_neto
                    new_k = compra_neto
                    new_g = new_c - old_c + new_j - new_k
                    matched = True
                    matched_pdf_keys.add(pdf_key)
                    break

        if not matched: continue

        if "DEUDA" not in nom_n:
            if pdf_costo_total is not None: new_b = pdf_costo_total
            else: new_b = max(0.0, old_b + new_k - new_j)
            actualizar_celda(ws, fila, 2, round(new_b, 2))
            old_b = new_b

        actualizar_celda(ws, fila, 3, round(new_c, 2))

        if old_b == 0 and new_c == 0:
            actualizar_celda(ws, fila, 5, 0.0)
            actualizar_celda(ws, fila, 6, 0.0)
            actualizar_celda(ws, fila, 7, 0.0)
            actualizar_celda(ws, fila, 8, 0.0)
        else:
            new_e = new_c - old_b
            actualizar_celda(ws, fila, 5, round(new_e, 2))
            new_f = (new_e / old_b) if old_b > 0 else 0.0
            actualizar_celda(ws, fila, 6, round(new_f, 10))
            if new_g is not None: actualizar_celda(ws, fila, 7, round(new_g, 2))
            if new_g is not None and old_b > 0:
                new_h = new_g / old_b
                actualizar_celda(ws, fila, 8, round(new_h, 10))

        actualizar_celda(ws, fila, 10, round(new_j, 2))
        actualizar_celda(ws, fila, 11, round(new_k, 2))
        actualizar_celda(ws, fila, 14, round(new_c, 2))

    if efectivo_instr and gbm:
        fila = efectivo_instr["fila"]
        old_c = valor_numerico(efectivo_instr["C"])
        gbm_total = gbm["valor_total"]
        sum_port = sum(d["valor"] for d in pdf_port.values())
        new_c = round(gbm_total - sum_port - deuda_gbm_matched, 2)
        if not deuda_sc_matched and deuda_sc_total > 0: new_c += deuda_sc_total
        if new_c < 0: new_c = 0.0

        actualizar_celda(ws, fila, 3, round(new_c, 2))
        actualizar_celda(ws, fila, 5, "-")
        actualizar_celda(ws, fila, 6, "-")
        actualizar_celda(ws, fila, 7, 0.0)
        actualizar_celda(ws, fila, 8, "-")
        actualizar_celda(ws, fila, 10, 0.0)
        actualizar_celda(ws, fila, 11, 0.0)
        actualizar_celda(ws, fila, 14, round(new_c, 2))

    nuevos = []
    for pdf_key, pdf_data in pdf_port.items():
        if pdf_key not in matched_pdf_keys:
            vm = pdf_data["valor"]
            ct = pdf_data.get("costo_total")
            compra = compras_map.get(pdf_key, 0)
            venta = ventas_map.get(pdf_key, 0)
            if vm == 0 and compra == 0 and venta == 0: continue
            nuevos.append({
                "emisora": pdf_key, "valor_a_mercado": vm, "costo_total": ct,
                "compra_neto": compra, "venta_neto": venta,
            })

    if nuevos:
        if efectivo_instr: fila_insercion = efectivo_instr["fila"]
        else: fila_insercion = fila_totales
        for nuevo in nuevos:
            fila_ref = fila_insercion - 1 if fila_insercion > fila_header + 1 else fila_header + 1
            insertar_instrumento(ws, fila_insercion, nuevo, fila_ref, periodo)
            fila_insercion += 1
            fila_totales += 1
        expandir_formulas_totales(ws, fila_totales)

    ultima_fila = ws.max_row
    if ultima_fila > fila_totales:
        ws.delete_rows(fila_totales + 1, ultima_fila - fila_totales)

    ultima_instr = fila_totales - 1
    _actualizar_formulas_header(ws, fila_header, fila_totales, ultima_instr)

    if efectivo_instr and gbm:
        efect_fila = None
        for r in range(fila_header + 1, fila_totales):
            nombre_r = ws.cell(r, 1).value
            if nombre_r and "EFECTIVO" in str(nombre_r).upper():
                efect_fila = r
                break
        if efect_fila:
            old_efect_b = valor_numerico(efectivo_instr["B"])
            total_compras = sum(compras_map.values())
            total_ventas = sum(ventas_map.values())
            efect_b = round(max(0.0, old_efect_b + total_ventas - total_compras), 2)
            ws.cell(efect_fila, 2).value = efect_b

    if periodo:
        actualizar_celda(ws, 2, 9, f"CORTE MENSUAL {periodo['mes_nombre']}", forzar=True)
        actualizar_celda(ws, 3, 9, periodo["periodo"], forzar=True)
        k7 = ws.cell(7, 11).value
        if k7 and "\n" in str(k7):
            actualizar_celda(ws, 7, 11, f"RENDIMIENTO ANUAL\n{periodo['anio']}", forzar=True)

def buscar_hoja(wb, nombre_cliente):
    nc = nombre_cliente.upper().strip()
    for s in wb.sheetnames:
        if s.upper().strip() == nc: return s
    for s in wb.sheetnames:
        if nc in s.upper() or s.upper().strip() in nc: return s
    partes = nc.split()
    for s in wb.sheetnames:
        su = s.upper()
        if sum(1 for p in partes if p in su) >= 2: return s
    return None

# ============================================================
# INTERFAZ STREAMLIT
# ============================================================
def main():
    st.title("🏦 Consolidador de Estados de Cuenta")
    
    col1, col2 = st.columns(2)
    with col1:
        maestro_file = st.file_uploader("📂 Sube el Maestro (.xlsx) Anterior", type=["xlsx"])
    with col2:
        pdf_files = st.file_uploader("📄 Sube los PDFs del Mes", type=["pdf"], accept_multiple_files=True)

    if st.button("🚀 Iniciar Consolidación", use_container_width=True):
        if maestro_file and pdf_files:
            try:
                # 1. Cargar archivo Excel
                wb = load_workbook(maestro_file)
                
                # 2. Procesar PDFs
                with st.spinner("Extrayendo datos de PDFs..."):
                    clientes = extraer_todos_los_pdfs_en_memoria(pdf_files)
                
                # 3. Consolidar
                with st.spinner("Actualizando archivo maestro..."):
                    clientes_actualizados = 0
                    for nombre_cliente, datos in sorted(clientes.items()):
                        hoja = buscar_hoja(wb, nombre_cliente)
                        if hoja:
                            actualizar_hoja(wb[hoja], datos, hoja)
                            clientes_actualizados += 1
                            st.success(f"✅ Hoja actualizada: {hoja}")
                        else:
                            st.warning(f"⚠️ Cliente no encontrado en el Maestro: {nombre_cliente}")
                
                # 4. Descargar
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                st.download_button(
                    label="📥 Descargar Estado de Cuenta Consolidado",
                    data=output,
                    file_name="ESTADOS_DE_CUENTA_ACTUALIZADOS.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            except Exception as e:
                import traceback
                st.error(f"❌ Error crítico: {e}")
                st.code(traceback.format_exc())
        else:
            st.error("Por favor, sube el Excel Maestro y al menos un PDF.")

if __name__ == "__main__":
    main()
